#!/usr/bin/python3

# https://pypi.org/
# pip install openpyxl         (install openpyxl package)

import openpyxl as xl
from openpyxl.chart import BarChart, Reference

def process_workbook(filename)
	wb = xl.load_workbook(filename)
	sheet = wb['Sheet1']
	cell = sheet['a1']          # access the perticular cell
	cell = sheet.cell(1,1)      # access the perticular cell
	print (cell.value)          # print cell value
	print (sheet.max_row)       # print row 

	for row in range (2, sheet.max_row + 1):  # 2,3,4 rows 
		print(row)
		cell = sheet.cell(row,3)
		print (cell.value)       # print values fron 3rd column 
		corrected_price = (cell.value * 0.9)       # corrected_price
		corrected_price_cell = sheet.cell(row,4)
		corrected_price_cell.value = corrected_price

	# create the instance of reference class and store in values object
	values= Reference(sheet,min_row=2, max_row = sheet.max_row,mix_col=4,max_cal=4)

	chart = BarChart()
	chart.add_data(values)
	sheet.add_chart(chart,"e2")   
		
	wb.save('transactions2.xlxs')    # save it in new file
	wb.save(filename)                # overwrite the file



